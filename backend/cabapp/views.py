from datetime import date, datetime, timedelta
import random
from django.db import transaction
from django.db.models import Count, Sum, Q
from django.http import HttpResponse
from django.conf import settings
from rest_framework import status, generics
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated, IsAdminUser
from rest_framework.response import Response
from rest_framework_simplejwt.tokens import RefreshToken
from rest_framework_simplejwt.views import TokenObtainPairView

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .models import Driver, Company, Ride, Attendance, PasswordResetOTP, Vehicle, CarCharge, AdvanceSalaryRequest
from .serializers import (
    CompanySerializer, RideSerializer, RideCreateSerializer,
    SyncPayloadSerializer, AttendanceSerializer, DriverSerializer,
    RegisterSerializer, UserSerializer, CabTokenObtainPairSerializer,
    ForgotPasswordRequestSerializer, VerifyOTPSerializer, ResetPasswordSerializer,
    CarChargeSerializer, VehicleSerializer, AdvanceSalaryRequestSerializer
)


def update_attendance(driver, ride_date):
    """Recalculate attendance for a driver on a given date."""
    total = Ride.objects.filter(driver=driver, date=ride_date).count()
    attendance, _ = Attendance.objects.get_or_create(driver=driver, date=ride_date)
    attendance.total_rides = total
    attendance.save()
    return attendance


class CabTokenObtainPairView(TokenObtainPairView):
    serializer_class = CabTokenObtainPairSerializer


@api_view(['POST'])
@permission_classes([])
def register(request):
    serializer = RegisterSerializer(data=request.data)
    if serializer.is_valid():
        user = serializer.save()
        refresh = RefreshToken.for_user(user)
        return Response({
            'user': UserSerializer(user).data,
            'access': str(refresh.access_token),
            'refresh': str(refresh),
        }, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


def _find_user_by_contact(contact):
    """Find a user by email or phone number."""
    contact = contact.strip()
    # Try email first
    try:
        from django.contrib.auth.models import User
        user = User.objects.get(email__iexact=contact)
        return user
    except User.DoesNotExist:
        pass
    # Try phone number via Driver model
    try:
        driver = Driver.objects.get(phone=contact)
        return driver.user
    except Driver.DoesNotExist:
        pass
    return None


@api_view(['POST'])
@permission_classes([])
def forgot_password(request):
    serializer = ForgotPasswordRequestSerializer(data=request.data)
    if not serializer.is_valid():
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    contact = serializer.validated_data['contact']
    user = _find_user_by_contact(contact)

    if not user:
        return Response(
            {'error': 'No account found with this phone number or email.'},
            status=status.HTTP_404_NOT_FOUND
        )

    # Invalidate any existing unused OTPs for this user
    PasswordResetOTP.objects.filter(user=user, is_used=False).update(is_used=True)

    # Generate 6-digit OTP
    otp_code = f"{random.randint(100000, 999999)}"
    PasswordResetOTP.objects.create(user=user, otp=otp_code)

    # Print OTP to console (replace with email/SMS in production)
    print(f"\n{'='*50}")
    print(f"  PASSWORD RESET OTP for {user.username}")
    print(f"  Contact: {contact}")
    print(f"  OTP: {otp_code}")
    print(f"{'='*50}\n")

    # Determine which method was used
    is_email = '@' in contact
    masked = _mask_contact(contact, is_email)

    return Response({
        'message': f'OTP sent to {masked}',
        'contact_type': 'email' if is_email else 'phone',
        'dev_otp': otp_code,
    })


def _mask_contact(contact, is_email):
    """Mask contact info for privacy."""
    if is_email:
        parts = contact.split('@')
        name = parts[0]
        masked_name = name[0] + '***' + (name[-1] if len(name) > 1 else '')
        return f"{masked_name}@{parts[1]}"
    else:
        return contact[:2] + '****' + contact[-2:] if len(contact) > 4 else '****'


@api_view(['POST'])
@permission_classes([])
def verify_otp(request):
    serializer = VerifyOTPSerializer(data=request.data)
    if not serializer.is_valid():
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    contact = serializer.validated_data['contact']
    otp = serializer.validated_data['otp']
    user = _find_user_by_contact(contact)

    if not user:
        return Response({'error': 'No account found.'}, status=status.HTTP_404_NOT_FOUND)

    otp_obj = PasswordResetOTP.objects.filter(user=user, otp=otp, is_used=False).first()
    if not otp_obj or not otp_obj.is_valid():
        return Response({'error': 'Invalid or expired OTP.'}, status=status.HTTP_400_BAD_REQUEST)

    return Response({'message': 'OTP verified successfully.'})


@api_view(['POST'])
@permission_classes([])
def reset_password(request):
    serializer = ResetPasswordSerializer(data=request.data)
    if not serializer.is_valid():
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    contact = serializer.validated_data['contact']
    otp = serializer.validated_data['otp']
    new_password = serializer.validated_data['new_password']
    user = _find_user_by_contact(contact)

    if not user:
        return Response({'error': 'No account found.'}, status=status.HTTP_404_NOT_FOUND)

    otp_obj = PasswordResetOTP.objects.filter(user=user, otp=otp, is_used=False).first()
    if not otp_obj or not otp_obj.is_valid():
        return Response({'error': 'Invalid or expired OTP.'}, status=status.HTTP_400_BAD_REQUEST)

    # Mark OTP as used and set new password
    otp_obj.is_used = True
    otp_obj.save()
    user.set_password(new_password)
    user.save()

    return Response({'message': 'Password reset successfully. You can now log in.'})


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def driver_dashboard(request):
    try:
        driver = request.user.driver_profile
    except Driver.DoesNotExist:
        return Response({'error': 'Driver profile not found'}, status=404)

    today = date.today()
    threshold = getattr(settings, 'FULL_DAY_THRESHOLD', 4)
    full_rate = getattr(settings, 'FULL_DAY_RATE', 800)
    half_rate = getattr(settings, 'HALF_DAY_RATE', 400)

    today_rides_qs = Ride.objects.filter(driver=driver, date=today)
    today_count = today_rides_qs.count()

    is_full = today_count >= threshold
    current_salary = full_rate if is_full else half_rate

    recent_rides = RideSerializer(
        today_rides_qs.select_related('company').order_by('-created_at')[:10],
        many=True
    ).data

    return Response({
        'driver_id': driver.id,
        'driver_name': driver.name,
        'today_rides': today_count,
        'target_rides': threshold,
        'status': 'Full Day' if is_full else 'Half Day',
        'salary_today': current_salary,
        'progress_percentage': min(100, (today_count / threshold) * 100),
        'recent_rides': recent_rides,
        'default_vehicle_number': driver.default_vehicle_number,
        'default_seater': driver.default_seater,
    })


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def create_ride(request):
    try:
        driver = request.user.driver_profile
    except Driver.DoesNotExist:
        return Response({'error': 'Driver profile not found'}, status=404)

    serializer = RideCreateSerializer(data=request.data, context={'request': request})
    if serializer.is_valid():
        # Check for duplicate local_id
        local_id = request.data.get('local_id')
        if local_id and Ride.objects.filter(local_id=local_id).exists():
            ride = Ride.objects.get(local_id=local_id)
            return Response(RideSerializer(ride).data, status=status.HTTP_200_OK)

        ride = serializer.save()
        update_attendance(driver, ride.date)
        return Response(RideSerializer(ride).data, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def create_car_charge(request):
    try:
        driver = request.user.driver_profile
    except Driver.DoesNotExist:
        return Response({'error': 'Driver profile not found'}, status=404)

    serializer = CarChargeSerializer(data=request.data, context={'request': request})
    if serializer.is_valid():
        charge = serializer.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def sync_rides(request):
    try:
        driver = request.user.driver_profile
    except Driver.DoesNotExist:
        return Response({'error': 'Driver profile not found'}, status=404)

    payload = SyncPayloadSerializer(data=request.data)
    if not payload.is_valid():
        return Response(payload.errors, status=status.HTTP_400_BAD_REQUEST)

    rides_data = payload.validated_data['rides']
    results = {'synced': [], 'skipped': [], 'errors': []}

    with transaction.atomic():
        for ride_data in rides_data:
            local_id = ride_data['local_id']

            if Ride.objects.filter(local_id=local_id).exists():
                results['skipped'].append(local_id)
                continue

            try:
                company = None
                if ride_data.get('company_id'):
                    company = Company.objects.get(id=ride_data['company_id'])

                ride = Ride.objects.create(
                    local_id=local_id,
                    driver=driver,
                    company=company,
                    date=ride_data['date'],
                    ride_time=ride_data.get('ride_time'),
                    trip_type=ride_data.get('trip_type', 'P'),
                    route=ride_data.get('route', ''),
                    pickup=ride_data.get('pickup', ''),
                    drop=ride_data.get('drop', ''),
                    notes=ride_data.get('notes', ''),
                    total_km=ride_data.get('total_km'),
                    vehicle_number=ride_data.get('vehicle_number', ''),
                )
                results['synced'].append(local_id)
                update_attendance(driver, ride.date)
            except Exception as e:
                results['errors'].append({'local_id': local_id, 'error': str(e)})

    return Response({
        'message': f"Sync complete. {len(results['synced'])} synced, {len(results['skipped'])} skipped.",
        **results
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def list_companies(request):
    companies = Company.objects.filter(is_active=True)
    return Response(CompanySerializer(companies, many=True).data)


@api_view(['POST'])
@permission_classes([IsAdminUser])
def create_company(request):
    serializer = CompanySerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['POST'])
@permission_classes([IsAdminUser])
def create_driver(request):
    serializer = RegisterSerializer(data=request.data)
    if serializer.is_valid():
        user = serializer.save()
        # We don't need to return tokens here as this is for admin to create drivers
        return Response({
            'user': UserSerializer(user).data,
            'driver': DriverSerializer(user.driver_profile).data
        }, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def admin_dashboard(request):
    if not (request.user.is_staff or request.user.is_superuser):
        return Response({'error': 'Admin access required'}, status=403)

    # Parse filters
    start_date = request.query_params.get('start_date')
    end_date = request.query_params.get('end_date')
    driver_id = request.query_params.get('driver_id')
    company_id = request.query_params.get('company_id')

    attendance_qs = Attendance.objects.all()
    ride_qs = Ride.objects.all()

    if start_date:
        attendance_qs = attendance_qs.filter(date__gte=start_date)
        ride_qs = ride_qs.filter(date__gte=start_date)
    if end_date:
        attendance_qs = attendance_qs.filter(date__lte=end_date)
        ride_qs = ride_qs.filter(date__lte=end_date)
    if driver_id:
        attendance_qs = attendance_qs.filter(driver_id=driver_id)
        ride_qs = ride_qs.filter(driver_id=driver_id)
    if company_id:
        ride_qs = ride_qs.filter(company_id=company_id)

    stats = attendance_qs.aggregate(
        total_half_days=Count('id', filter=Q(status='Half')),
        total_full_days=Count('id', filter=Q(status='Full')),
        total_salary=Sum('salary'),
    )

    # Advance salary stats
    advance_qs = AdvanceSalaryRequest.objects.select_related('driver').all()
    if driver_id:
        advance_qs = advance_qs.filter(driver_id=driver_id)
    pending_advance_count = advance_qs.filter(status='Pending').count()

    # Paid advances with driver names
    paid_advances = advance_qs.filter(status='Paid')
    total_advance_paid = sum(float(a.amount) for a in paid_advances)
    advance_paid_drivers = list(set(a.driver.name for a in paid_advances))

    return Response({
        'total_rides': ride_qs.count(),
        'total_drivers': Driver.objects.filter(is_active=True).count(),
        'total_half_days': stats['total_half_days'] or 0,
        'total_full_days': stats['total_full_days'] or 0,
        'total_salary_paid': stats['total_salary'] or 0,
        'pending_advance_count': pending_advance_count,
        'total_advance_paid': total_advance_paid,
        'advance_paid_drivers': advance_paid_drivers,
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def reports(request):
    if not (request.user.is_staff or request.user.is_superuser):
        return Response({'error': 'Admin access required'}, status=403)

    start_date = request.query_params.get('start_date')
    end_date = request.query_params.get('end_date')
    driver_id = request.query_params.get('driver_id')
    company_id = request.query_params.get('company_id')

    qs = Attendance.objects.select_related('driver').all()

    if start_date:
        qs = qs.filter(date__gte=start_date)
    if end_date:
        qs = qs.filter(date__lte=end_date)
    if driver_id:
        qs = qs.filter(driver_id=driver_id)
    if company_id:
        driver_ids_with_company = Ride.objects.filter(company_id=company_id)
        if start_date:
            driver_ids_with_company = driver_ids_with_company.filter(date__gte=start_date)
        if end_date:
            driver_ids_with_company = driver_ids_with_company.filter(date__lte=end_date)
        driver_ids = driver_ids_with_company.values_list('driver_id', flat=True).distinct()
        qs = qs.filter(driver_id__in=driver_ids)

    results = []
    for att in qs.order_by('-date', 'driver__name'):
        rides_qs = Ride.objects.select_related('company').filter(driver=att.driver, date=att.date)
        if company_id:
            rides_qs = rides_qs.filter(company_id=company_id)
        rides = RideSerializer(rides_qs.order_by('ride_time', 'created_at'), many=True).data
        results.append({
            'driver_name': att.driver.name,
            'driver_id': att.driver_id,
            'date': att.date,
            'total_rides': att.total_rides,
            'status': att.status,
            'salary': att.salary,
            'rides': rides,
        })

    return Response(results)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_excel(request):
    if not (request.user.is_staff or request.user.is_superuser):
        return Response({'error': 'Admin access required'}, status=403)

    start_date_str = request.query_params.get('start_date')
    end_date_str = request.query_params.get('end_date')
    driver_id = request.query_params.get('driver_id')
    company_id = request.query_params.get('company_id')

    # Parse date range
    try:
        d_start = date.fromisoformat(start_date_str) if start_date_str else date.today().replace(day=1)
        d_end = date.fromisoformat(end_date_str) if end_date_str else date.today()
    except ValueError:
        d_start = date.today().replace(day=1)
        d_end = date.today()

    # Generate all dates in range
    all_dates = []
    current = d_start
    while current <= d_end:
        all_dates.append(current)
        current += timedelta(days=1)

    # Get drivers to include
    if driver_id:
        target_drivers = list(Driver.objects.filter(id=driver_id, is_active=True))
    else:
        target_drivers = list(Driver.objects.filter(is_active=True).order_by('name'))

    # Fetch attendance data
    att_qs = Attendance.objects.select_related('driver').filter(
        date__gte=d_start, date__lte=d_end
    )
    if driver_id:
        att_qs = att_qs.filter(driver_id=driver_id)

    # Build attendance lookup: {(driver_id, date): attendance}
    att_lookup = {}
    for att in att_qs:
        att_lookup[(att.driver_id, att.date)] = att

    # Fetch charges for the period
    charge_qs_all = CarCharge.objects.filter(date__gte=d_start, date__lte=d_end)
    if driver_id:
        charge_qs_all = charge_qs_all.filter(driver_id=driver_id)

    # Fetch advance salary requests (Paid) for the period
    advance_qs_all = AdvanceSalaryRequest.objects.filter(
        status='Paid',
        request_date__date__gte=d_start,
        request_date__date__lte=d_end,
    )
    if driver_id:
        advance_qs_all = advance_qs_all.filter(driver_id=driver_id)

    full_rate = getattr(settings, 'FULL_DAY_RATE', 1000)
    half_rate = getattr(settings, 'HALF_DAY_RATE', 500)
    threshold = getattr(settings, 'FULL_DAY_THRESHOLD', 4)

    wb = openpyxl.Workbook()

    header_fill = PatternFill(start_color='1F3A5F', end_color='1F3A5F', fill_type='solid')
    charge_header_fill = PatternFill(start_color='D32F2F', end_color='D32F2F', fill_type='solid')
    advance_header_fill = PatternFill(start_color='1565C0', end_color='1565C0', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    center_align = Alignment(horizontal='center')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    full_fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
    half_fill = PatternFill(start_color='FFF9C4', end_color='FFF9C4', fill_type='solid')
    leave_fill = PatternFill(start_color='FFCDD2', end_color='FFCDD2', fill_type='solid')
    holiday_fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')

    first = True
    for drv in target_drivers:
        # 1. Rides Sheet
        sheet_name = drv.name[:31]
        if first:
            ws = wb.active
            ws.title = sheet_name
            first = False
        else:
            ws = wb.create_sheet(title=sheet_name)

        # Title
        ws.merge_cells('A1:H1')
        title_cell = ws['A1']
        title_cell.value = f"Ride Manifest - {drv.name}"
        title_cell.font = Font(bold=True, size=13)
        title_cell.alignment = center_align

        # Headers
        headers = ['Date', 'No', 'Client', 'P/D', 'Time', 'Route', 'Total km', 'Status']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border

        total_salary = 0
        total_rides_count = 0
        row_idx = 3

        for d in all_dates:
            att = att_lookup.get((drv.id, d))
            is_weekend = d.weekday() in (5, 6)

            if att and att.total_rides > 0:
                # Has rides - show ride details
                rides_qs = Ride.objects.select_related('company').filter(driver=drv, date=d)
                if company_id:
                    rides_qs = rides_qs.filter(company_id=company_id)
                rides = list(rides_qs.order_by('ride_time', 'created_at')) or [None]

                day_status = att.status
                day_salary = float(att.salary)
                total_rides_count += att.total_rides

                for ride_number, ride in enumerate(rides, 1):
                    if ride:
                        route = ride.route or f"{ride.pickup} to {ride.drop}"
                        data = [
                            d.strftime('%d-%m-%Y'),
                            ride_number,
                            ride.company.name if ride.company else '',
                            ride.trip_type,
                            ride.ride_time.strftime('%I:%M %p').lstrip('0') if ride.ride_time else '',
                            route,
                            float(ride.total_km) if ride.total_km is not None else '',
                            day_status if ride_number == 1 else '',
                        ]
                    else:
                        data = [d.strftime('%d-%m-%Y'), '', '', '', '', '', '', day_status]

                    row_fill = full_fill if day_status == 'Full' else half_fill
                    for col, val in enumerate(data, 1):
                        cell = ws.cell(row=row_idx, column=col, value=val)
                        cell.alignment = center_align if col != 6 else Alignment(horizontal='left')
                        cell.border = border
                        cell.fill = row_fill
                    row_idx += 1
                total_salary += day_salary
            else:
                # No rides: Leave or Holiday
                if is_weekend:
                    day_label = 'Holiday'
                    row_fill = holiday_fill
                else:
                    day_label = 'Leave'
                    row_fill = leave_fill

                data = [d.strftime('%d-%m-%Y'), '', '', '', '', '', '', day_label]
                for col, val in enumerate(data, 1):
                    cell = ws.cell(row=row_idx, column=col, value=val)
                    cell.alignment = center_align
                    cell.border = border
                    cell.fill = row_fill
                row_idx += 1

        # Rides Total Summary
        total_row = row_idx + 1
        ws.cell(row=total_row, column=1, value='TOTAL RIDES').font = Font(bold=True)
        ws.cell(row=total_row, column=2, value=total_rides_count).font = Font(bold=True)
        ws.cell(row=total_row, column=6, value='TOTAL SALARY').font = Font(bold=True)
        ws.cell(row=total_row, column=8, value=total_salary).font = Font(bold=True)

        # 2. Charges Sheet
        charge_sheet_name = f"{drv.name[:20]} - Charges"
        wsc = wb.create_sheet(title=charge_sheet_name)
        
        # Title
        wsc.merge_cells('A1:H1')
        title_cell_c = wsc['A1']
        title_cell_c.value = f"Car Charge Details - {drv.name}"
        title_cell_c.font = Font(bold=True, size=13, color='FFFFFF')
        title_cell_c.fill = charge_header_fill
        title_cell_c.alignment = center_align
        
        # Headers
        charge_headers = ['Date', 'App Used', 'Time', 'Place', 'Vehicle', '', '', 'Amount']
        for col, header in enumerate(charge_headers, 1):
            if header:
                cell = wsc.cell(row=2, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
                cell.border = border
        
        driver_charges = charge_qs_all.filter(driver=drv).order_by('date', 'time')
        
        total_charges = 0
        crow_idx = 3
        for charge in driver_charges:
            data = [
                charge.date.strftime('%d-%m-%Y'),
                charge.app_used,
                charge.time.strftime('%I:%M %p').lstrip('0'),
                charge.place,
                charge.vehicle_number,
                None,
                None,
                float(charge.charge_amount)
            ]
            for col, val in enumerate(data, 1):
                if val is not None:
                    cell = wsc.cell(row=crow_idx, column=col, value=val)
                    cell.alignment = center_align
                    cell.border = border
            total_charges += float(charge.charge_amount)
            crow_idx += 1
            
        # Charge Total
        total_charges_row = crow_idx + 1
        wsc.cell(row=total_charges_row, column=1, value='TOTAL CHARGES').font = Font(bold=True)
        wsc.cell(row=total_charges_row, column=8, value=total_charges).font = Font(bold=True)

        # 3. Advance Salary on Rides Sheet
        driver_advances = advance_qs_all.filter(driver=drv)
        total_advance = sum(float(a.amount) for a in driver_advances)
        
        # 4. Final Summary on Rides Sheet
        summary_row = total_row + 2
        ws.merge_cells(f'A{summary_row}:H{summary_row}')
        summary_header = ws.cell(row=summary_row, column=1, value="PAYMENT SUMMARY")
        summary_header.font = Font(bold=True, size=11, color='FFFFFF')
        summary_header.fill = header_fill
        summary_header.alignment = center_align
        
        row_idx = summary_row + 1
        ws.cell(row=row_idx, column=1, value='Total Salary (A)').font = Font(bold=True)
        ws.cell(row=row_idx, column=8, value=total_salary)

        row_idx += 1
        ws.cell(row=row_idx, column=1, value='Advance Salary Paid (B)').font = Font(bold=True)
        adv_cell = ws.cell(row=row_idx, column=8, value=total_advance)
        adv_cell.font = Font(bold=True, color='1565C0')
        
        row_idx += 1
        ws.cell(row=row_idx, column=1, value='SALARY TO BE PAID (A - B)').font = Font(bold=True)
        net_cell = ws.cell(row=row_idx, column=8, value=total_salary - total_advance)
        net_cell.font = Font(bold=True, size=12, color='2E7D32')
        net_cell.border = Border(bottom=Side(style='double'))

        # Column widths for both sheets
        col_widths = [15, 15, 20, 8, 14, 34, 12, 14]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
            wsc.column_dimensions[get_column_letter(i)].width = width

    # Consolidated Charges Sheet (Only if charges exist)
    if charge_qs_all.exists():
        wsall = wb.create_sheet(title="All Car Charges", index=0)
        
        # Title
        wsall.merge_cells('A1:G1')
        title_cell_a = wsall['A1']
        title_cell_a.value = "Consolidated Car Charges - All Drivers"
        title_cell_a.font = Font(bold=True, size=14, color='FFFFFF')
        title_cell_a.fill = charge_header_fill
        title_cell_a.alignment = center_align
        
        # Headers
        all_headers = ['Date', 'Driver Name', 'App Used', 'Time', 'Place', 'Vehicle', 'Amount']
        for col, header in enumerate(all_headers, 1):
            cell = wsall.cell(row=2, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border
            
        row_idx_a = 3
        total_all_charges = 0
        for charge in charge_qs_all.select_related('driver').order_by('date', 'time'):
            data = [
                charge.date.strftime('%d-%m-%Y'),
                charge.driver.name,
                charge.app_used,
                charge.time.strftime('%I:%M %p').lstrip('0'),
                charge.place,
                charge.vehicle_number,
                float(charge.charge_amount)
            ]
            for col, val in enumerate(data, 1):
                cell = wsall.cell(row=row_idx_a, column=col, value=val)
                cell.alignment = center_align
                cell.border = border
            total_all_charges += float(charge.charge_amount)
            row_idx_a += 1
            
        # Total row
        row_idx_a += 1
        wsall.cell(row=row_idx_a, column=1, value='GRAND TOTAL').font = Font(bold=True)
        wsall.cell(row=row_idx_a, column=7, value=total_all_charges).font = Font(bold=True)
        
        # Widths
        col_widths_a = [15, 20, 20, 14, 30, 15, 15]
        for i, width in enumerate(col_widths_a, 1):
            wsall.column_dimensions[get_column_letter(i)].width = width

    if not target_drivers:
        ws = wb.active
        ws.title = 'Report'
        ws['A1'] = 'No data found for the selected filters.'

    filename = f"ride_manifest"
    if start_date_str:
        filename += f"_{start_date_str}"
    if end_date_str:
        filename += f"_to_{end_date_str}"
    filename += ".xlsx"

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_driver_excel(request):
    try:
        driver = request.user.driver_profile
    except Driver.DoesNotExist:
        return Response({'error': 'Driver profile not found'}, status=403)

    period = request.query_params.get('period')
    today = date.today()
    
    qs = Attendance.objects.filter(driver=driver)
    
    if period == 'month':
        qs = qs.filter(date__year=today.year, date__month=today.month)
        period_str = today.strftime('%b_%Y')
    elif period == 'year':
        qs = qs.filter(date__year=today.year)
        period_str = str(today.year)
    else:
        period_str = 'all_time'

    qs = qs.order_by('date')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{driver.name[:20]}_{period_str}"

    header_fill = PatternFill(start_color='1F3A5F', end_color='1F3A5F', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    center_align = Alignment(horizontal='center')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    ws.merge_cells('A1:H1')
    title_cell = ws['A1']
    title_cell.value = f"Ride Manifest - {driver.name} ({period_str.replace('_', ' ')})"
    title_cell.font = Font(bold=True, size=13)
    title_cell.alignment = center_align

    headers = ['Date', 'No', 'Client', 'P/D', 'Time', 'Route', 'Total km', 'Vehicle']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border

    total_salary = 0
    row_idx = 3
    attendances = list(qs)
    
    for att in attendances:
        rides = list(Ride.objects.select_related('company').filter(driver=driver, date=att.date).order_by('ride_time', 'created_at')) or [None]

        for ride_number, ride in enumerate(rides, 1):
            if ride:
                route = ride.route or f"{ride.pickup} to {ride.drop}"
                data = [
                    att.date.strftime('%d-%m-%Y'),
                    ride_number,
                    ride.company.name if ride.company else '',
                    ride.trip_type,
                    ride.ride_time.strftime('%I:%M %p').lstrip('0') if ride.ride_time else '',
                    route,
                    float(ride.total_km) if ride.total_km is not None else '',
                    ride.vehicle_number,
                ]
            else:
                data = [att.date.strftime('%d-%m-%Y'), '', '', '', '', '', '', '']

            for col, val in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.alignment = center_align if col != 6 else Alignment(horizontal='left')
                cell.border = border
                if att.status == 'Full':
                    cell.fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
                else:
                    cell.fill = PatternFill(start_color='FFF9C4', end_color='FFF9C4', fill_type='solid')
            row_idx += 1
        total_salary += float(att.salary)

    total_row = row_idx + 1
    ws.cell(row=total_row, column=1, value='TOTAL RIDES').font = Font(bold=True)
    ws.cell(row=total_row, column=2, value=sum(a.total_rides for a in attendances)).font = Font(bold=True)
    salary_cell = ws.cell(row=total_row, column=6, value=f"Salary: {total_salary}")
    salary_cell.font = Font(bold=True)

    col_widths = [15, 8, 20, 8, 14, 34, 12, 14]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    if not attendances:
        ws = wb.active
        ws['A1'] = f"No rides found for {period_str}."

    filename = f"{driver.name.replace(' ', '_')}_Report_{period_str}.xlsx"

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def list_drivers(request):
    if not (request.user.is_staff or request.user.is_superuser):
        return Response({'error': 'Admin access required'}, status=403)
    drivers = Driver.objects.filter(is_active=True).select_related('user')
    return Response(DriverSerializer(drivers, many=True).data)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def driver_report_data(request):
    try:
        driver = request.user.driver_profile
    except Driver.DoesNotExist:
        return Response({'error': 'Driver profile not found'}, status=403)

    period = request.query_params.get('period')
    today = date.today()
    
    qs = Attendance.objects.filter(driver=driver)
    if period == 'month':
        qs = qs.filter(date__year=today.year, date__month=today.month)
    elif period == 'year':
        qs = qs.filter(date__year=today.year)
    
    qs = qs.order_by('date')
    
    results = []
    for att in qs:
        rides_qs = Ride.objects.select_related('company').filter(driver=driver, date=att.date).order_by('ride_time', 'created_at')
        rides = RideSerializer(rides_qs, many=True).data
        results.append({
            'date': att.date,
            'status': att.status,
            'total_rides': att.total_rides,
            'salary': att.salary,
            'rides': rides
        })
        
    return Response({
        'driver_name': driver.name,
        'period': period,
        'data': results
    })

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def list_vehicles(request):
    vehicles = Vehicle.objects.filter(is_active=True)
    return Response(VehicleSerializer(vehicles, many=True).data)


@api_view(['POST'])
@permission_classes([IsAdminUser])
def create_vehicle(request):
    serializer = VehicleSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def update_default_vehicle(request):
    try:
        driver = request.user.driver_profile
    except Driver.DoesNotExist:
        return Response({'error': 'Driver profile not found'}, status=404)

    number = request.data.get('default_vehicle_number', '')
    seater = request.data.get('default_seater', 4)

    driver.default_vehicle_number = number
    driver.default_seater = seater
    driver.save()

    return Response({'message': 'Default vehicle updated successfully'})


# ──────────── Advance Salary Endpoints ────────────

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def request_advance_salary(request):
    """Driver requests an advance salary."""
    try:
        driver = request.user.driver_profile
    except Driver.DoesNotExist:
        return Response({'error': 'Driver profile not found'}, status=404)

    amount = request.data.get('amount')
    reason = request.data.get('reason', '')

    if not amount or float(amount) <= 0:
        return Response({'error': 'Please enter a valid amount.'}, status=400)

    advance = AdvanceSalaryRequest.objects.create(
        driver=driver,
        amount=amount,
        reason=reason,
    )
    return Response(
        AdvanceSalaryRequestSerializer(advance).data,
        status=status.HTTP_201_CREATED
    )


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def list_advance_requests(request):
    """Admin lists all advance salary requests. Drivers see only their own."""
    if request.user.is_staff or request.user.is_superuser:
        qs = AdvanceSalaryRequest.objects.select_related('driver').all()
        status_filter = request.query_params.get('status')
        if status_filter:
            qs = qs.filter(status=status_filter)
    else:
        try:
            driver = request.user.driver_profile
        except Driver.DoesNotExist:
            return Response({'error': 'Driver profile not found'}, status=404)
        qs = AdvanceSalaryRequest.objects.filter(driver=driver)

    return Response(AdvanceSalaryRequestSerializer(qs, many=True).data)


@api_view(['PATCH'])
@permission_classes([IsAdminUser])
def update_advance_request(request, pk):
    """Admin updates an advance salary request status."""
    try:
        advance = AdvanceSalaryRequest.objects.get(pk=pk)
    except AdvanceSalaryRequest.DoesNotExist:
        return Response({'error': 'Request not found'}, status=404)

    new_status = request.data.get('status')
    if new_status not in ('Paid', 'Rejected'):
        return Response({'error': 'Status must be Paid or Rejected.'}, status=400)

    advance.status = new_status
    advance.resolved_date = datetime.now()
    advance.save()

    return Response(AdvanceSalaryRequestSerializer(advance).data)

